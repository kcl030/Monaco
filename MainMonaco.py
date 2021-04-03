# -*- coding: utf-8 -*-
"""
Created on Sat Dec 26 15:52:17 2020

@author: Iris
"""

from TestingMScreen import mainReadScreen
import os
import xlrd
import xlwt
from xlwt import Workbook
from openpyxl import load_workbook
from MonacoSort import list_difference



def MapEntries(directory,Mapindex,DeletePics=False):

    directory = r'C:\Program Files (x86)\Steam\userdata\82949594\760\remote\113020\screenshots'
    
    
    FinalClock = []
    FinalName = []
    FinalSec = []
    
    
    
    class Entryoutput:
            def __init__(self,name, clock,rank,mapname,maptype):
                self.name = name
                self.clock = clock
                self.rank = rank
                self.mapname = mapname
                self.maptype = maptype
                
    
    
    #loops through every picture
    for filename in os.listdir(directory):
        if filename.endswith(".jpg") or filename.endswith(".png"):
            
            fullpath = os.path.join(directory, filename)
            print(fullpath)
            
            #Grabs Name and clock from ReadMonacoScreenShot   
            ScreenOutput = mainReadScreen(fullpath)
            
            subName = ScreenOutput.name
            subName = subName.replace('\x0c', '')   #remove return
            
            subNamelist = subName.split("\n")
            subNamelist = list(filter(None, subNamelist))
          
            thing = ' '  
            while thing in subNamelist: subNamelist.remove(thing)   #removes nothing
            
          
            
            subClock = ScreenOutput.clock
            subClock = subClock.replace('\x0c', '')  #remove return
            subClock = subClock.replace('.', '')     #remove trailing decimals
            subClock = subClock.replace('O', '0')     #replace o with 0
            subClock = subClock.replace(',', '')     #remove trailing decimals
            
            
            subClocklist = subClock.split("\n")
            subClocklist = list(filter(None, subClocklist))   
         
            subSec = ScreenOutput.milisec   
         
            
            #Builds final leaderboard
            FinalClock.extend(subClocklist)
            FinalName.extend(subNamelist)
            FinalSec.extend(subSec)
            
            #DELETES THE PICTURE. 
            if DeletePics == True:
                os.remove(fullpath)
            
            
        else:
            continue
    
    
    
    
    
    ##################################
    #Need to implement:  Read Excel
    ##################################
    
    wbpath = r'C:\Program Files (x86)\Steam\userdata\82949594\760\remote\113020\screenshots\MonacoScreenshots.xlsx'
    
    #Open Workboook
    wb = xlrd.open_workbook(wbpath)
    sheet = wb.sheet_by_index(Mapindex)
    
    #Read book
    
    OldNames = []
    OldClock = []
    
    for x in range(4,54):
        OldNames.append(sheet.cell_value(x, 1))
        OldClock.append(sheet.cell_value(x, 2))
    
    
    
    #Compare to  Final parameters
    ClockEntrysIndex =  [i for i, item in enumerate(FinalClock) if not item in set(OldClock)]
    
    
    FinalEntry = []
    for x in range(len(ClockEntrysIndex)):
    
        EntryName = FinalName[ClockEntrysIndex[x]]
        EntryClock= FinalClock[ClockEntrysIndex[x]]
        
    
        Entry = Entryoutput(EntryName, EntryClock , (ClockEntrysIndex[x]+1),sheet.cell_value(1,1),sheet.cell_value(1,0) )
    
        FinalEntry.append(Entry)
    
    
    
    
    
    #################
    #Write to Excel
    #################
    
    wb2 = load_workbook(wbpath)
    a = wb2.sheetnames[Mapindex]
    
    sheeta = wb2[a]
    
    for x in range(4,54):
         
         sheeta.cell(x+1,2).value  = str(FinalName[x-4])
         sheeta.cell(x+1,3).value  = str(FinalClock[x-4])
         #sheeta.cell(x+1,4).value  = str(FinalSec[x-4])
    
    wb2.save(wbpath)

